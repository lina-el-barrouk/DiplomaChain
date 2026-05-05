"""
DiplomaChain – Endpoints PDF
POST /pdf/templates/upload     → institution uploade son template
POST /pdf/templates/positions  → configurer les positions des champs
GET  /pdf/diplomas/{id}/generate → générer le PDF d'un diplôme
GET  /pdf/templates/preview    → prévisualiser le template
"""
import io
import json
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.v1.deps import get_current_user, require_approved_institution
from app.db.database import get_db
from app.models.models import Diploma, DiplomaTemplate, Institution, Student, User
# APRÈS
from app.services.pdf_service import (
    generate_diploma_pdf,
    save_generated_pdf,
    save_template_file,
)
from app.core.security import decrypt_sensitive

router = APIRouter(prefix="/pdf", tags=["PDF Diplômes"])

MAX_TEMPLATE_SIZE = 20 * 1024 * 1024  # 20 Mo


# ─────────────────────────────────────────────────────────────────────────────
# SCHEMAS
# ─────────────────────────────────────────────────────────────────────────────

class FieldPosition(BaseModel):
    x: float
    y: float
    font: str = "Helvetica"
    font_size: float = 12
    color: str = "#000000"
    align: str = "center"


class QRPosition(BaseModel):
    x: float
    y: float
    width: float = 90
    height: float = 90


class TemplatePositions(BaseModel):
    student_name: FieldPosition | None = None
    degree_title: FieldPosition | None = None
    field_of_study: FieldPosition | None = None
    institution_name: FieldPosition | None = None
    graduation_date: FieldPosition | None = None
    honors: FieldPosition | None = None
    unique_code: FieldPosition | None = None
    qr_code: QRPosition | None = None


# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD DU TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/templates/upload")
async def upload_template(
    file: UploadFile = File(..., description="Template PDF du diplôme (max 20 Mo)"),
    current_user: User = Depends(require_approved_institution),
    db: Session = Depends(get_db),
):
    """
    L'institution uploade son design de diplôme vierge en PDF.
    Le backend stocke le fichier et crée/met à jour le template.
    """
    if file.content_type != "application/pdf":
        raise HTTPException(400, "Seuls les fichiers PDF sont acceptés")

    file_bytes = await file.read()
    if len(file_bytes) > MAX_TEMPLATE_SIZE:
        raise HTTPException(413, "Fichier trop volumineux (max 20 Mo)")

    institution = db.query(Institution).filter(
        Institution.manager_id == current_user.id
    ).first()

    # Sauvegarder le fichier
    file_path = save_template_file(file_bytes, institution.id, file.filename)

    # Créer ou mettre à jour le template en base
    template = db.query(DiplomaTemplate).filter(
        DiplomaTemplate.institution_id == institution.id
    ).first()

    if template:
        template.file_name = file.filename
        template.file_path = file_path
        template.is_active = True
    else:
        template = DiplomaTemplate(
            institution_id=institution.id,
            file_name=file.filename,
            file_path=file_path,
            field_positions="{}",
        )
        db.add(template)

    db.commit()
    db.refresh(template)

    return {
        "message": "Template uploadé avec succès",
        "template_id": template.id,
        "file_name": template.file_name,
        "positions": json.loads(template.field_positions),
    }


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURER LES POSITIONS
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/templates/positions")
def update_positions(
    positions: TemplatePositions,
    current_user: User = Depends(require_approved_institution),
    db: Session = Depends(get_db),
):
    """
    L'institution configure les coordonnées X/Y de chaque champ sur son template.
    """
    institution = db.query(Institution).filter(
        Institution.manager_id == current_user.id
    ).first()

    template = db.query(DiplomaTemplate).filter(
        DiplomaTemplate.institution_id == institution.id
    ).first()

    if not template:
        raise HTTPException(404, "Aucun template trouvé — uploadez d'abord un template")

    # Fusionner avec les positions existantes
    existing = json.loads(template.field_positions)
    new_positions = positions.model_dump(exclude_none=True)

    for field, pos in new_positions.items():
        existing[field] = pos

    template.field_positions = json.dumps(existing)
    db.commit()

    return {
        "message": "Positions mises à jour",
        "positions": existing,
    }


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE EXCEL POUR GÉNÉRATION EN MASSE
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/diplomas/bulk-template")
def download_bulk_template(
    current_user: User = Depends(require_approved_institution),
):
    """
    Télécharge le fichier Excel modèle (.xlsx) à remplir pour la génération en masse.
    Colonnes : code_massar | titre_diplome | domaine | date_graduation | mention
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Diplômes"

    COLUMNS = [
        ("code_massar",  "Code Massar de l'étudiant (doit exister dans le système)"),
        ("titre_diplome",   "Titre du diplôme (ex : Master en Informatique)"),
        ("domaine",         "Domaine d'études (ex : Intelligence Artificielle)"),
        ("date_graduation", "Date de graduation (AAAA-MM-JJ ou JJ/MM/AAAA)"),
        ("mention",         "Mention : Passable / Assez Bien / Bien / Très Bien / Excellent (optionnel)"),
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # En-têtes
    for col_idx, (col_key, col_desc) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_key)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = 30

    # Ligne d'exemple
    ws.append([
        "P123456789",
        "Master en Informatique",
        "Intelligence Artificielle",
        "2024-06-30",
        "Très Bien",
    ])
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_diplomes_masse.xlsx"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GÉNÉRATION EN MASSE DEPUIS UN FICHIER EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/diplomas/bulk-generate")
async def bulk_generate_diplomas(
    file: UploadFile = File(..., description="Fichier Excel .xlsx (max 10 Mo)"),
    current_user: User = Depends(require_approved_institution),
    db: Session = Depends(get_db),
):
    """
    Génère les diplômes en masse depuis un fichier Excel.

    Pour chaque ligne valide :
      1. Cherche l'étudiant par email (doit être enregistré et approuvé par l'admin)
      2. Crée le diplôme en statut PENDING (pas d'émission automatique)
      3. Génère un PDF de prévisualisation

    L'institution doit ensuite émettre chaque diplôme manuellement depuis
    le tableau de bord — ce qui déclenche l'ancrage sur Hedera HCS.

    Retourne un fichier ZIP contenant :
      - Un PDF de prévisualisation par diplôme : diplome_<email>_<code>.pdf
      - Un rapport : rapport.csv
    """
    import csv
    import zipfile
    from datetime import datetime, timezone

    from openpyxl import load_workbook

    from app.core.security import decrypt_sensitive
    from app.models.models import User as UserModel
    from app.schemas.diploma import DiplomaCreate
    from app.services.diploma_service import create_diploma, generate_unique_code

    # ── Validation du fichier ──────────────────────────────────────────────
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Seuls les fichiers Excel (.xlsx / .xls) sont acceptés")

    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(413, "Fichier trop volumineux (max 10 Mo)")

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f"Impossible de lire le fichier Excel : {exc}")

    # ── Lire les en-têtes ─────────────────────────────────────────────────
    raw_headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
    REQUIRED = {"code_massar", "titre_diplome", "domaine", "date_graduation"}
    missing = REQUIRED - set(raw_headers)
    if missing:
        raise HTTPException(400, f"Colonnes manquantes dans l'Excel : {', '.join(sorted(missing))}")

    col = {h: i for i, h in enumerate(raw_headers)}

    # ── Récupérer l'institution et son template ───────────────────────────
    institution = db.query(Institution).filter(
        Institution.manager_id == current_user.id
    ).first()

    from app.models.models import DiplomaTemplate
    tmpl = db.query(DiplomaTemplate).filter(
        DiplomaTemplate.institution_id == institution.id,
        DiplomaTemplate.is_active == True,
    ).first()
    template_path = tmpl.file_path if tmpl else None

    # ── Traiter chaque ligne ───────────────────────────────────────────────
    results: list[dict] = []
    pdf_files: dict[str, bytes] = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Ignorer les lignes vides
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(col_name: str):
            idx = col.get(col_name)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        massar_code    = cell("code_massar")
        degree_title   = cell("titre_diplome")
        field_of_study = cell("domaine")
        grad_raw       = cell("date_graduation")
        honors         = cell("mention") or None

        def err(msg: str):
            results.append({"ligne": row_num, "code_massar": massar_code or "", "statut": "erreur",
                             "code": "", "message": msg})

        # Validation des champs obligatoires
        if not massar_code or not degree_title or not field_of_study or not grad_raw:
            err("Champs obligatoires manquants (code_massar, titre, domaine, date)")
            continue

        # Parsing de la date
        try:
            raw_val = row[col["date_graduation"]]
            if isinstance(raw_val, datetime):
                graduation_date = raw_val.replace(tzinfo=timezone.utc)
            else:
                date_str = str(raw_val).strip()
                parsed = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                    try:
                        parsed = datetime.strptime(date_str, fmt).replace(tzinfo=timezone.utc)
                        break
                    except ValueError:
                        pass
                if parsed is None:
                    raise ValueError(f"Format inconnu : {date_str}")
                graduation_date = parsed
        except Exception as exc:
            err(f"Date invalide : {exc}")
            continue

        # Chercher l'étudiant
        import hashlib
        massar_code_hash = hashlib.sha256(massar_code.encode()).hexdigest()
        student = db.query(Student).filter(Student.massar_code_hash == massar_code_hash).first()
        
        if not student:
            err("Aucun étudiant trouvé avec ce Code Massar")
            continue

        if not student.is_approved:
            err("L'étudiant n'a pas encore été approuvé par l'administrateur")
            continue

        # ── ÉTAPE 1 : Créer le diplôme en PENDING (commit immédiat) ──────────
        try:
            diploma_schema = DiplomaCreate(
                student_id=student.id,
                degree_title=degree_title,
                field_of_study=field_of_study,
                graduation_date=graduation_date,
                honors=honors,
            )
            diploma = create_diploma(db=db, data=diploma_schema, institution_id=institution.id)
            # create_diploma fait déjà db.commit() — le diplôme est persisté en 'pending'
        except Exception as exc:
            db.rollback()
            err(f"Impossible de créer le diplôme : {exc}")
            continue  # Passer à la ligne suivante

        # ── ÉTAPE 2 : Générer le PDF de prévisualisation (échec n'annule pas le diplôme) ──
        try:
            student_name = decrypt_sensitive(student.full_name_enc)
            pdf_data = {
                "degree_title":  diploma.degree_title,
                "field_of_study": diploma.field_of_study,
                "graduation_date": diploma.graduation_date,
                "honors":        diploma.honors,
                "unique_code":   diploma.unique_code,
            }
            pdf_bytes = generate_diploma_pdf(
                diploma_data=pdf_data,
                student_name=student_name,
                institution_name=institution.name,
                template_path=template_path,
            )
            save_generated_pdf(pdf_bytes, diploma.id)

            safe_massar = massar_code.replace(" ", "_")
            filename = f"diplome_{safe_massar}_{diploma.unique_code}.pdf"
            pdf_files[filename] = pdf_bytes

            results.append({
                "ligne": row_num, "code_massar": massar_code,
                "statut": "succès", "code": diploma.unique_code,
                "message": "Diplôme créé (en attente d'émission par l'institution)",
            })

        except Exception as pdf_exc:
            # Le diplôme est déjà en base — on note juste que le PDF a échoué
            results.append({
                "ligne": row_num, "code_massar": massar_code,
                "statut": "succès", "code": diploma.unique_code,
                "message": f"Diplôme créé (PDF indisponible : {pdf_exc})",
            })

    # ── Vérifier si au moins un diplôme a été créé ───────────────────────
    diplomas_created = sum(1 for r in results if r.get("statut") == "succès")
    if diplomas_created == 0:
        raise HTTPException(
            400,
            detail={
                "message": "Aucun diplôme n'a pu être créé. Consultez le rapport ci-dessous.",
                "rapport": results,
            },
        )

    # ── Construire le ZIP (PDFs + rapport CSV) ────────────────────────────
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, pdf_b in pdf_files.items():
            zf.writestr(fname, pdf_b)

        # Rapport CSV
        csv_io = io.StringIO()
        writer = csv.DictWriter(
            csv_io,
            fieldnames=["ligne", "email", "statut", "code", "message"],
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(results)
        zf.writestr("rapport.csv", csv_io.getvalue())

    zip_buf.seek(0)

    success_n = diplomas_created
    total_n   = len(results)

    return Response(
        content=zip_buf.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=diplomes_masse_{institution.id[:8]}.zip",
            "X-Success-Count": str(success_n),
            "X-Total-Count":   str(total_n),
            "Access-Control-Expose-Headers": "X-Success-Count, X-Total-Count",
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# GÉNÉRER LE PDF D'UN DIPLÔME
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/diplomas/{diploma_id}/generate")
def generate_diploma(
    diploma_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Génère le PDF final du diplôme avec :
    - Le template de l'institution (ou fond par défaut)
    - Les infos de l'étudiant
    - Le QR code inséré automatiquement
    Retourne le PDF en téléchargement.
    """
    # Récupérer le diplôme
    diploma = db.query(Diploma).filter(Diploma.id == diploma_id).first()
    if not diploma:
        raise HTTPException(404, "Diplôme introuvable")

    if diploma.status != "issued":
        raise HTTPException(400, "Le diplôme doit être émis avant de générer le PDF")

    # Contrôle d'accès
    if current_user.role == "institution":
        institution = db.query(Institution).filter(
            Institution.manager_id == current_user.id
        ).first()
        if not institution or diploma.institution_id != institution.id:
            raise HTTPException(403, "Accès refusé")
    elif current_user.role == "student":
        student = db.query(Student).filter(Student.user_id == current_user.id).first()
        if not student or diploma.student_id != student.id:
            raise HTTPException(403, "Accès refusé")
    elif current_user.role != "admin":
        raise HTTPException(403, "Accès refusé")

    # Récupérer les infos de l'étudiant (déchiffrement)
    student = db.query(Student).filter(Student.id == diploma.student_id).first()
    student_name = decrypt_sensitive(student.full_name_enc) if student else "Étudiant"

    # Récupérer l'institution
    institution = db.query(Institution).filter(
        Institution.id == diploma.institution_id
    ).first()
    institution_name = institution.name if institution else "Institution"

    # Récupérer le template de l'institution
    template = db.query(DiplomaTemplate).filter(
        DiplomaTemplate.institution_id == diploma.institution_id,
        DiplomaTemplate.is_active == True,
    ).first()

    template_path = template.file_path if template else None
    field_positions = json.loads(template.field_positions) if template else None

    # Générer le PDF
    diploma_data = {
        "degree_title": diploma.degree_title,
        "field_of_study": diploma.field_of_study,
        "graduation_date": diploma.graduation_date,
        "honors": diploma.honors,
        "unique_code": diploma.unique_code,
    }

    pdf_bytes = generate_diploma_pdf(
        diploma_data=diploma_data,
        student_name=student_name,
        institution_name=institution_name,
        template_path=template_path,
        field_positions=field_positions,
    )

    # Sauvegarder une copie
    save_generated_pdf(pdf_bytes, diploma.id)

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=diplome-{diploma.unique_code}.pdf",
            "Cache-Control": "no-store",
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# VOIR MON TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/templates/me")
def get_my_template(
    current_user: User = Depends(require_approved_institution),
    db: Session = Depends(get_db),
):
    """Retourne les infos du template de l'institution."""
    institution = db.query(Institution).filter(
        Institution.manager_id == current_user.id
    ).first()

    template = db.query(DiplomaTemplate).filter(
        DiplomaTemplate.institution_id == institution.id
    ).first()

    if not template:
        return {
            "has_template": False,
            "message": "Aucun template configuré — un fond par défaut sera utilisé",
            "message": "Aucun template — un fond élégant par défaut sera utilisé avec placeholders",
        }

    return {
        "has_template": True,
        "template_id": template.id,
        "file_name": template.file_name,
        "is_active": template.is_active,
        "positions": json.loads(template.field_positions),
        "created_at": template.created_at,
    }
